import io
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from app.auth.dependencies import get_current_user
from app.database import get_db
from app.models.auth_models import User, UserRole
from app.models.models import ValidationRun


def _get_run_for_user(run_id: int, db: Session, user: User) -> ValidationRun:
    q = db.query(ValidationRun).filter(ValidationRun.id == run_id)
    if user.role != UserRole.RHADIX_ADMIN:
        q = q.filter(ValidationRun.tenant_id == user.tenant_id)
    run = q.first()
    if not run:
        raise HTTPException(404, "Run not found")
    return run

router = APIRouter()

SEV_LABEL = {"error": "Fout", "warning": "Waarschuwing", "info": "Info"}

@router.get("/{run_id}/excel")
def export_excel(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    run = _get_run_for_user(run_id, db, current_user)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = openpyxl.Workbook()

    # ── Samenvatting sheet ──
    ws = wb.active
    ws.title = "Samenvatting"
    dark  = "0A0F1E"
    blue  = "6366F1"
    teal  = "0EA5E9"
    red   = "EF4444"
    amber = "F59E0B"
    green = "22C55E"
    white = "F1F5F9"
    gray  = "94A3B8"

    def hdr_cell(ws, row, col, value, bg=blue, fg=white, bold=True, size=11):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return c

    def data_cell(ws, row, col, value, fg="1E293B", bold=False, bg=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(color=fg, bold=bold, size=10, name="Calibri")
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return c

    thin = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Title
    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value=f"Rhadix Validatierapport — {run.label}")
    t.font = Font(bold=True, size=14, color=white, name="Calibri")
    t.fill = PatternFill("solid", fgColor=dark)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.cell(row=2, column=1, value=f"Aangemaakt: {run.created_at.strftime('%d-%m-%Y %H:%M')}").font = Font(size=9, color=gray, name="Calibri")
    ws.row_dimensions[2].height = 18

    # Stats row
    stats = [("Bestanden", len(run.files or [])), ("Rijen", run.total_rows), ("Fouten", run.error_count), ("Waarschuwingen", run.warn_count), ("Score", f"{run.score}%")]
    row = 4
    for i, (lbl, val) in enumerate(stats, 1):
        hdr_cell(ws, row, i, lbl, bg="1E293B", fg=gray, size=9)
        c = data_cell(ws, row+1, i, str(val), fg=red if (lbl=="Fouten" and run.error_count>0) else amber if (lbl=="Waarschuwingen" and run.warn_count>0) else green if lbl=="Score" else white, bold=True)
        c.font = Font(bold=True, size=13, color=red if (lbl=="Fouten" and run.error_count>0) else amber if (lbl=="Waarschuwingen" and run.warn_count>0) else green if lbl=="Score" else "0EA5E9", name="Calibri")
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row+1].height = 28

    # Per-file summary table
    row = 7
    for col, hdr in enumerate(["Bestand","Schema","Rijen","Fouten","Waarschuwingen"],1):
        hdr_cell(ws, row, col, hdr, bg="1E293B", size=10)
    ws.row_dimensions[row].height = 22
    results = run.results or {}
    for fr in results.get("file_results", []):
        row += 1
        data_cell(ws, row, 1, fr["filename"])
        data_cell(ws, row, 2, fr["schema_key"])
        data_cell(ws, row, 3, fr["row_count"])
        data_cell(ws, row, 4, fr["error_count"], fg=red if fr["error_count"]>0 else "22C55E", bold=fr["error_count"]>0)
        data_cell(ws, row, 5, fr["warn_count"],  fg=amber if fr["warn_count"]>0 else "22C55E",  bold=fr["warn_count"]>0)
        for c in range(1,6): ws.cell(row=row, column=c).border = thin
        ws.row_dimensions[row].height = 18

    for i, w in enumerate([38,20,10,10,14],1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Issues sheet ──
    wi = wb.create_sheet("Bevindingen")
    wi.freeze_panes = "A2"
    hdrs = ["Bestand","Soort","Ernst","Omschrijving","Aantal","Details"]
    for col, hdr in enumerate(hdrs, 1):
        hdr_cell(wi, 1, col, hdr, bg=dark, size=10)
    wi.row_dimensions[1].height = 22

    SEV_COLOR = {"error": red, "warning": amber, "info": blue}
    row = 1
    for fr in results.get("file_results", []):
        for issue in fr.get("issues", []):
            row += 1
            data_cell(wi, row, 1, fr["filename"])
            data_cell(wi, row, 2, "Bestand")
            sev = issue["severity"]
            c = wi.cell(row=row, column=3, value=SEV_LABEL.get(sev, sev))
            c.font = Font(color=SEV_COLOR.get(sev, blue), bold=True, size=10, name="Calibri")
            data_cell(wi, row, 4, issue["label"])
            data_cell(wi, row, 5, issue["count"])
            data_cell(wi, row, 6, issue.get("detail",""))
            wi.row_dimensions[row].height = 16

    for issue in results.get("cross_results", []):
        row += 1
        data_cell(wi, row, 1, "—")
        data_cell(wi, row, 2, "Cross-bestand")
        sev = issue["severity"]
        c = wi.cell(row=row, column=3, value=SEV_LABEL.get(sev, sev))
        c.font = Font(color=SEV_COLOR.get(sev, blue), bold=True, size=10, name="Calibri")
        data_cell(wi, row, 4, issue["label"])
        data_cell(wi, row, 5, issue["count"])
        data_cell(wi, row, 6, issue.get("detail",""))
        wi.row_dimensions[row].height = 16

    for i, w in enumerate([32,16,14,42,10,40],1):
        wi.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Rhadix_Rapport_{run.id}_{run.created_at.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/{run_id}/pdf")
def export_pdf(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    run = _get_run_for_user(run_id, db, current_user)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
    except ImportError:
        raise HTTPException(500, "reportlab not installed")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=18*mm, bottomMargin=18*mm)

    C_DARK  = colors.HexColor("#0A0F1E")
    C_BLUE  = colors.HexColor("#6366F1")
    C_TEAL  = colors.HexColor("#0EA5E9")
    C_RED   = colors.HexColor("#EF4444")
    C_AMBER = colors.HexColor("#F59E0B")
    C_GREEN = colors.HexColor("#22C55E")
    C_GRAY  = colors.HexColor("#64748B")
    C_LIGHT = colors.HexColor("#F1F5F9")
    C_BG    = colors.HexColor("#0F172A")

    styles = getSampleStyleSheet()
    def style(name, **kw):
        return ParagraphStyle(name, fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#1E293B"), **kw)

    title_style  = style("title",  fontSize=18, textColor=C_DARK, fontName="Helvetica-Bold", spaceAfter=2)
    sub_style    = style("sub",    fontSize=9,  textColor=C_GRAY, spaceAfter=12)
    h2_style     = style("h2",     fontSize=11, textColor=C_DARK, fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    body_style   = style("body",   fontSize=9,  textColor=colors.HexColor("#334155"))

    results = run.results or {}
    story = []

    story.append(Paragraph("Rhadix Validatierapport", title_style))
    story.append(Paragraph(f"{run.label} · {run.created_at.strftime('%d %B %Y, %H:%M')}", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=C_BLUE, spaceAfter=14))

    # Stats table
    stats_data = [["Bestanden","Rijen","Fouten","Waarschuwingen","Score"],
                  [str(len(run.files or [])), str(run.total_rows), str(run.error_count), str(run.warn_count), f"{run.score}%"]]
    st = Table(stats_data, colWidths=[34*mm]*5)
    st.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), C_BG),
        ("TEXTCOLOR",  (0,0), (-1,0), C_GRAY),
        ("BACKGROUND", (0,1), (-1,1), C_LIGHT),
        ("FONTNAME",   (0,0), (-1,-1), "Helvetica"),
        ("FONTNAME",   (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 8),
        ("FONTSIZE",   (0,1), (-1,1), 13),
        ("TEXTCOLOR",  (2,1), (2,1), C_RED if run.error_count>0 else C_GREEN),
        ("TEXTCOLOR",  (3,1), (3,1), C_AMBER if run.warn_count>0 else C_GREEN),
        ("TEXTCOLOR",  (4,1), (4,1), C_GREEN),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,0),(-1,-1), [None]),
        ("BOX",        (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ("INNERGRID",  (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
    ]))
    story.append(st)
    story.append(Spacer(1, 14))

    # Per-file results
    story.append(Paragraph("Resultaten per bestand", h2_style))
    for fr in results.get("file_results", []):
        schema_label = fr["schema_key"].capitalize()
        story.append(Paragraph(f"<b>{fr['filename']}</b> ({schema_label} · {fr['row_count']} rijen)", body_style))
        if not fr.get("issues"):
            story.append(Paragraph("✓ Geen bevindingen", ParagraphStyle("ok", fontName="Helvetica", fontSize=9, textColor=C_GREEN, leftIndent=8, spaceAfter=6)))
        else:
            issue_data = [["Ernst","Omschrijving","Aantal","Details"]]
            for iss in fr["issues"]:
                issue_data.append([
                    SEV_LABEL.get(iss["severity"],iss["severity"]),
                    iss["label"], str(iss["count"]), iss.get("detail","") or ""
                ])
            it = Table(issue_data, colWidths=[18*mm, 60*mm, 14*mm, 68*mm])
            sev_colors = {"Fout": C_RED, "Waarschuwing": C_AMBER, "Info": C_BLUE}
            ts = [
                ("BACKGROUND",(0,0),(-1,0),C_BG),("TEXTCOLOR",(0,0),(-1,0),C_GRAY),
                ("FONTNAME",(0,0),(-1,0),"Helvetica"),("FONTSIZE",(0,0),(-1,-1),8),
                ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#F8FAFC"),colors.white]),
                ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
                ("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ]
            for i, iss in enumerate(fr["issues"],1):
                col = sev_colors.get(SEV_LABEL.get(iss["severity"],""), C_BLUE)
                ts.append(("TEXTCOLOR",(0,i),(0,i),col))
                ts.append(("FONTNAME",(0,i),(0,i),"Helvetica-Bold"))
            it.setStyle(TableStyle(ts))
            story.append(it)
        story.append(Spacer(1,8))

    # Cross results
    if results.get("cross_results"):
        story.append(Paragraph("Cross-bestand controles", h2_style))
        cross_data = [["Ernst","Omschrijving","Aantal","Details"]]
        for c in results["cross_results"]:
            cross_data.append([SEV_LABEL.get(c["severity"],c["severity"]), c["label"], str(c["count"]), c.get("detail","") or ""])
        ct = Table(cross_data, colWidths=[18*mm,65*mm,14*mm,63*mm])
        ct.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),C_BG),("TEXTCOLOR",(0,0),(-1,0),C_GRAY),
            ("FONTNAME",(0,0),(-1,0),"Helvetica"),("FONTSIZE",(0,0),(-1,-1),8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#F8FAFC"),colors.white]),
            ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
            ("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ]))
        story.append(ct)

    doc.build(story)
    buf.seek(0)
    filename = f"Rhadix_Rapport_{run.id}_{run.created_at.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ─── Actieplan PDF export ──────────────────────────────────────────────────────

class EstimateInfo(BaseModel):
    label: str = ""
    minutes: int = 0
    issueCount: int = 0
    rowCount: int = 0
    breakdown: str = ""

class ActieplanItem(BaseModel):
    title: str
    color: str = "amber"
    desc: str = ""
    acties: list[str] = []
    estimate: EstimateInfo = EstimateInfo()

class ActieplanExportRequest(BaseModel):
    items: list[ActieplanItem]
    run_id: Optional[int] = None
    organisation: Optional[str] = None


def _format_time(minutes: int) -> str:
    if minutes < 60:
        return f"{minutes} min"
    h = minutes // 60
    m = minutes % 60
    return f"~{h}u {m}min" if m else f"~{h}u"


@router.post("/actieplan")
def export_actieplan(
    req: ActieplanExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, KeepTogether,
        )
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    except ImportError:
        raise HTTPException(500, "reportlab not installed")

    # ── Kleuren (zelfde palet als de rest van Rhadix) ──
    C_DARK  = colors.HexColor("#0A0F1E")
    C_BLUE  = colors.HexColor("#6366F1")
    C_TEAL  = colors.HexColor("#0EA5E9")
    C_RED   = colors.HexColor("#EF4444")
    C_AMBER = colors.HexColor("#F59E0B")
    C_GREEN = colors.HexColor("#22C55E")
    C_GRAY  = colors.HexColor("#64748B")
    C_LIGHT = colors.HexColor("#F1F5F9")
    C_BG    = colors.HexColor("#0F172A")
    C_WHITE = colors.white

    def ps(name, **kw):
        return ParagraphStyle(name, fontName="Helvetica", fontSize=9,
                              textColor=colors.HexColor("#1E293B"), **kw)

    sty_title   = ps("t",  fontSize=20, textColor=C_WHITE,  fontName="Helvetica-Bold", spaceAfter=0)
    sty_sub     = ps("s",  fontSize=9,  textColor=C_GRAY,   spaceAfter=0)
    sty_h2      = ps("h2", fontSize=11, textColor=C_DARK,   fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=5)
    sty_body    = ps("b",  fontSize=9,  textColor=colors.HexColor("#334155"), leading=14)
    sty_caption = ps("c",  fontSize=8,  textColor=C_GRAY,   leading=12)
    sty_bullet  = ps("bl", fontSize=9,  textColor=colors.HexColor("#334155"), leading=13,
                     leftIndent=12, bulletIndent=4, bulletText="•")
    sty_note    = ps("n",  fontSize=8,  textColor=C_TEAL,   leading=12,
                     fontName="Helvetica-Oblique")
    sty_footer  = ps("f",  fontSize=8,  textColor=C_GRAY,   alignment=TA_CENTER)

    # ── Optionele scandata ──
    run = None
    if req.run_id:
        run = db.query(ValidationRun).filter(ValidationRun.id == req.run_id).first()

    now_str      = datetime.now().strftime("%d-%m-%Y")
    scan_label   = run.label if run else "—"
    scan_date    = run.created_at.strftime("%d %B %Y") if run else now_str
    score        = run.score if run else None
    score2       = max(0, score - 26) if score is not None else None
    rhadix_index = round((score + score2) / 2) if score is not None else None

    # ── Tijdtotaal ──
    total_minutes = sum(item.estimate.minutes for item in req.items) or 30
    total_issues  = sum(item.estimate.issueCount for item in req.items)
    total_rows    = sum(item.estimate.rowCount   for item in req.items)

    time_parts = []
    if total_issues: time_parts.append(f"{total_issues} issues × 5 min")
    if total_rows:   time_parts.append(f"{total_rows} rijen × 0,5 min")
    if req.items:    time_parts.append(f"{len(req.items)} × 15 min controle")
    total_breakdown = " + ".join(time_parts) if time_parts else "schatting op basis van ernstniveau"

    impact = "Hoog" if len(req.items) >= 2 else ("Gemiddeld" if len(req.items) == 1 else "Laag")

    # ── Document ──
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm,
    )

    story = []

    # ── Titelbalk ──
    title_data = [[
        Paragraph("Rhadix Actieplan", sty_title),
        Paragraph(f"{now_str}", ps("tr", fontSize=9, textColor=C_GRAY, alignment=TA_RIGHT)),
    ]]
    title_table = Table(title_data, colWidths=[120*mm, 50*mm])
    title_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), C_BG),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("LEFTPADDING",   (0, 0), (0, -1),  12),
        ("RIGHTPADDING",  (-1, 0), (-1, -1), 12),
    ]))
    story.append(title_table)
    story.append(Spacer(1, 2))

    # ── Scan-info balk (grijs) ──
    org_str  = req.organisation or "Rhadix Validator"
    idx_str  = f"Rhadix Index: {rhadix_index}/100" if rhadix_index is not None else ""
    scan_str = f"Scan: {scan_label} · {scan_date}"
    info_parts = [x for x in [org_str, scan_str, idx_str] if x]
    info_row = [[Paragraph("  ·  ".join(info_parts), ps("ir", fontSize=8, textColor=C_GRAY))]]
    info_table = Table(info_row, colWidths=[170*mm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#F1F5F9")),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING",   (0, 0), (-1, -1), 12),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 12))

    # ── Samenvattingstabel ──
    story.append(Paragraph("Samenvatting", sty_h2))
    summary_headers = ["Acties", "Geschatte tijd", "Breakdown", "Impact"]
    summary_values  = [
        str(len(req.items)),
        _format_time(total_minutes),
        total_breakdown,
        impact,
    ]
    summ_table = Table(
        [summary_headers, summary_values],
        colWidths=[20*mm, 28*mm, 90*mm, 22*mm],
    )
    summ_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), C_BG),
        ("TEXTCOLOR",     (0, 0), (-1, 0), C_GRAY),
        ("BACKGROUND",    (0, 1), (-1, 1), C_LIGHT),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica"),
        ("FONTNAME",      (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 8),
        ("FONTSIZE",      (0, 1), (-1, 1), 10),
        ("TEXTCOLOR",     (1, 1), (1, 1), C_BLUE),
        ("TEXTCOLOR",     (3, 1), (3, 1), C_GREEN),
        ("ALIGN",         (0, 0), (1, -1), "CENTER"),
        ("ALIGN",         (2, 0), (2, -1), "LEFT"),
        ("ALIGN",         (3, 0), (3, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("BOX",           (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ]))
    story.append(summ_table)
    story.append(Spacer(1, 16))

    # ── Per actie ──
    story.append(Paragraph("Actiepunten", sty_h2))
    story.append(HRFlowable(width="100%", thickness=1, color=C_BLUE, spaceAfter=10))

    # Sortering: errors (rood) eerst
    sorted_items = sorted(req.items, key=lambda x: 0 if x.color == "red" else 1)

    for idx, item in enumerate(sorted_items):
        is_error   = item.color == "red"
        bar_color  = C_RED if is_error else C_AMBER
        prio_label = "Hoge prioriteit" if is_error else "Gemiddelde prioriteit"
        prio_color = colors.HexColor("#FEF2F2") if is_error else colors.HexColor("#FFFBEB")
        est        = item.estimate
        tijd_str   = est.label if est.label else _format_time(est.minutes or 30)

        # Prioriteitsbadge + titel
        header_data = [[
            Paragraph(f"{'✕' if is_error else '⚠'}  {item.title}",
                      ps(f"h{idx}", fontSize=11, textColor=C_DARK, fontName="Helvetica-Bold")),
            Paragraph(prio_label,
                      ps(f"p{idx}", fontSize=8, textColor=bar_color,
                         alignment=TA_RIGHT, fontName="Helvetica-Bold")),
        ]]
        header_table = Table(header_data, colWidths=[120*mm, 50*mm])
        header_table.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), prio_color),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (0, -1),  10),
            ("RIGHTPADDING",  (-1, 0), (-1, -1), 10),
            ("LINEAFTER",     (0, 0), (0, -1),  0, C_WHITE),
            ("LINEBEFORE",    (0, 0), (0, -1),  3, bar_color),
        ]))

        # Beschrijving
        desc_para = Paragraph(item.desc, sty_body) if item.desc else None

        # Aanbevolen acties
        actie_paras = [Paragraph(a, sty_bullet) for a in item.acties] if item.acties else []

        # Tijdschatting toelichting
        breakdown_str = est.breakdown or "Schatting op basis van ernstniveau"
        tijd_para = Paragraph(
            f"⏱ Tijdschatting: <b>{tijd_str}</b> — {breakdown_str}",
            sty_note,
        )

        block = [header_table]
        if desc_para:
            block.append(Spacer(1, 6))
            block.append(desc_para)
        if actie_paras:
            block.append(Spacer(1, 6))
            block.append(Paragraph("Aanbevolen acties:", ps(f"al{idx}", fontSize=9,
                                   fontName="Helvetica-Bold", textColor=C_DARK)))
            block.extend(actie_paras)
        block.append(Spacer(1, 6))
        block.append(tijd_para)
        block.append(Spacer(1, 12))

        story.append(KeepTogether(block))

    # ── Lege staat ──
    if not req.items:
        story.append(Paragraph("Geen actiepunten geselecteerd.", sty_caption))
        story.append(Spacer(1, 12))

    # ── Footer ──
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_GRAY, spaceBefore=16, spaceAfter=8))
    story.append(Paragraph(
        f"Gegenereerd door Rhadix Validator · {now_str}",
        sty_footer,
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"Rhadix_Actieplan_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
