import io
import csv
import json
import xml.etree.ElementTree as ET
import re
import os
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import List, Optional
from app.auth.dependencies import get_optional_user, require_app_access
from app.database import get_db
from app.services import run_cache
from app.models.auth_models import Application, TenantApplication, User, UserApplication
from app.models.models import ValidationRun
from app.services.validator import validate_files, detect_schema, auto_map, KIKV_REFERENCE, cross_checks_from_files
from app.services.concept_validator import validate_concept_mapping
from app.services.zib_validator import validate_zib
from app.services.zib_rules import detect_zib_schema
from app.services.algemeen_validator import validate_algemeen, _detect_template
from app.services.algemeen_benchmark import benchmark_against_reference
from app.services.actuality_validator import validate_actuality, detect_date_fields, get_kikv_norm_for_schema
from app.services.traceability import enrich_file_result, collect_all_issues
from app.services.owl_validator import validate_structural, validate_relational
from app.services.sparql_validator import validate_use_cases

router = APIRouter()

MAX_ROWS = int(os.getenv("RHADIX_MAX_ROWS", "100000"))  # harde bovengrens per bestand; ruim hoog zodat normale HR-exports niet stilletjes worden afgekapt

# Mapping from 'standard' form-param to Application slug
_STANDARD_TO_APP_SLUG = {
    "kikv":     "kikv-validator",
    "zib":      "zib-validator",
    "algemeen": "algemeen-validator",
}


def _resolve_app_and_license(standard: str, user, db):
    """
    Return (application_id, license_id) for the given standard and user.
    Both are None for anonymous (demo) users.
    """
    if not user:
        return None, None
    slug = _STANDARD_TO_APP_SLUG.get(standard)
    if not slug:
        return None, None
    from sqlalchemy.orm import Session
    app = db.query(Application).filter(Application.slug == slug, Application.is_active == True).first()
    if not app:
        return None, None
    # Find the UserApplication → TenantApplication → License chain
    ua = db.query(UserApplication).filter(
        UserApplication.user_id == user.id,
        UserApplication.application_id == app.id,
    ).first()
    if not ua:
        return app.id, None
    ta = ua.tenant_application
    return app.id, (ta.license_id if ta else None)


def parse_csv_bytes(content: bytes, filename: str, max_rows: int = MAX_ROWS) -> tuple[list, list]:
    text = content.decode("utf-8-sig", errors="replace")
    first_line = text.split("\n")[0]
    delim = ";" if first_line.count(";") > first_line.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    headers = reader.fieldnames or []
    rows = []
    total = 0
    for row in reader:
        clean = {k.strip('"').strip(): v.strip('"').strip() for k, v in row.items() if k}
        if any(v for v in clean.values()):
            total += 1
            if len(rows) < max_rows:
                rows.append(clean)
    return list(headers), rows, total


def _normalize_xml_value(val: str) -> str:
    """Strip tijdzone/tijd van datetime-strings en normaliseer YYYYMMDD naar dd-mm-yyyy."""
    if not val:
        return val
    # ISO datetime: 2026-04-19T00:00:00 → 2026-04-19
    val = re.sub(r'T\d{2}:\d{2}:\d{2}.*$', '', val.strip())
    # YYYYMMDD zonder streepjes → yyyy-mm-dd zodat datumparser het herkent
    if re.fullmatch(r'\d{8}', val):
        val = f"{val[:4]}-{val[4:6]}-{val[6:]}"
    return val


def parse_xml_bytes(content: bytes, max_rows: int = MAX_ROWS) -> tuple[list, list]:
    """Parseer AFAS Profit XML-exports (Profit_Employees, Profit_Timetable, Profit_Illness)."""
    try:
        root = ET.fromstring(content.decode("utf-8", errors="replace"))
    except ET.ParseError as e:
        raise ValueError(f"Ongeldig XML-bestand: {e}")

    # AFAS kent twee exportvormen: named export (<Profit_Employees><Employee>) en
    # GET-connector (<root><skip/><take/><rows><row>…). Kies de juiste container.
    rows_el = root.find("rows")
    container = rows_el if (rows_el is not None and len(list(rows_el)) > 0) else root

    headers_ordered: list[str] = []
    headers_seen: set[str] = set()
    rows: list[dict] = []
    total = 0

    for record in container:
        row: dict[str, str] = {}
        for field in record:
            tag = field.tag
            is_nil = (field.get("nil") or "").lower() == "true"
            val = "" if is_nil else _normalize_xml_value(field.text or "")
            row[tag] = val
            if tag not in headers_seen:
                headers_seen.add(tag)
                headers_ordered.append(tag)
        if any(v for v in row.values()):
            total += 1
            if len(rows) < max_rows:
                rows.append(row)

    return headers_ordered, rows, total


def parse_json_bytes(content: bytes, max_rows: int = MAX_ROWS) -> tuple[list, list]:
    """Parseer AFAS GetConnector JSON-exports.

    Ondersteunt de REST GetConnector-envelope {"skip":..,"take":..,"rows":[..]},
    een losse record-dict, en een platte lijst van records. Waarden worden
    gestringificeerd en datums genormaliseerd, identiek aan de XML-parser, zodat
    XML- en JSON-import dezelfde rijen opleveren. null -> "".
    """
    text = content.decode("utf-8-sig", errors="replace")
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # AFAS Profit exporteert getallen soms met voorloopnul (bijv.
        # "HouseNumber": 00), wat strikt genomen ongeldige JSON is. Repareer
        # alleen die ongequote getal-tokens en probeer opnieuw; blijft het kapot,
        # dan geven we een nette fout.
        try:
            data = json.loads(re.sub(r'(:\s*)0+(\d+)(\s*[,}\]])', r'\1\2\3', text))
        except json.JSONDecodeError as e:
            raise ValueError(f"Ongeldig JSON-bestand: {e}")

    if isinstance(data, dict):
        records = data.get("rows")
        if records is None:
            records = [data]            # losse record-dict -> één rij
    elif isinstance(data, list):
        records = data
    else:
        raise ValueError("JSON moet een object met 'rows' of een lijst van records zijn")

    headers_ordered: list[str] = []
    headers_seen: set[str] = set()
    rows: list[dict] = []
    total = 0

    for record in records:
        if not isinstance(record, dict):
            continue
        row: dict[str, str] = {}
        for key, raw in record.items():
            if raw is None:                       # null -> ""
                val = ""
            elif isinstance(raw, bool):           # bool -> "True"/"False" (AFAS-XML-casing)
                val = str(raw)
            elif isinstance(raw, (int, float)):
                # Numeriek getypeerde AFAS-waarde: wél stringificeren, maar NIET door
                # de datum-normalisatie halen — anders wordt bijv. een 8-cijferig
                # personeelsnummer/BSN onterecht als YYYYMMDD-datum geïnterpreteerd.
                val = str(raw)
            else:
                val = _normalize_xml_value(str(raw))
            row[key] = val
            if key not in headers_seen:
                headers_seen.add(key)
                headers_ordered.append(key)
        if any(v for v in row.values()):
            total += 1
            if len(rows) < max_rows:
                rows.append(row)

    return headers_ordered, rows, total


def parse_upload(content: bytes, filename: str, ext: str) -> tuple[list, list, int]:
    if ext == "csv":
        return parse_csv_bytes(content, filename)
    if ext == "xml":
        return parse_xml_bytes(content)
    if ext == "json":
        return parse_json_bytes(content)
    try:
        import openpyxl
        try:
            # read_only=True gebruikt de streaming-reader die de volledige
            # stylesheet niet inlaadt; omzeilt o.a. AFAS/Excel-exports die
            # falen met "expected <class 'openpyxl.styles.fills.Fill'>".
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        headers = [str(c or "") for c in all_rows[0]] if all_rows else []
        rows = [
            {headers[i]: str(cell or "") for i, cell in enumerate(row)}
            for row in all_rows[1:]
            if any(cell for cell in row)
        ]
        return headers, rows, len(rows)
    except Exception as e:
        raise HTTPException(400, f"Could not parse Excel file: {e}")


@router.post("/upload")
async def upload_and_validate(
    files: List[UploadFile] = File(...),
    label: Optional[str] = Form(None),
    standard: Optional[str] = Form("kikv"),   # "kikv" | "zib" | "algemeen"
    source: Optional[str] = Form(None),       # bron: "afas" | "ons" | "kikv_csv" | "epd_ecd"
    max_age_days: Optional[int] = Form(30),   # drempel voor actualiteitscheck
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_optional_user),
):
    """
    Upload bestanden en valideer tegen KIK-V (standard='kikv') of ZIB's (standard='zib').
    Voegt altijd 'actuality' toe aan de response.
    """
    standard = (standard or "kikv").lower().strip()

    # Bron-gestuurde flow: de opgegeven bron bepaalt de fase-1 standaard
    SOURCE_TO_STANDARD = {"afas": "algemeen", "ons": "algemeen",
                          "kikv_csv": "kikv", "kikv": "kikv",
                          "epd_ecd": "zib", "zib": "zib"}
    if source:
        source = source.lower().strip()
        standard = SOURCE_TO_STANDARD.get(source, standard)

    # ── Phase 2 app-level access check ──────────────────────────────────────
    # Authenticated users must have the relevant application assigned.
    # Anonymous (demo) users bypass this check.
    if current_user is not None:
        from app.auth.dependencies import require_app_access as _rac
        from app.models.auth_models import UserApplication, Application as _App
        from fastapi import HTTPException as _HTTPEx
        slug = _STANDARD_TO_APP_SLUG.get(standard)
        if slug:
            _app = db.query(_App).filter(_App.slug == slug, _App.is_active == True).first()
            from app.models.auth_models import UserRole
            if _app and current_user.role != UserRole.RHADIX_ADMIN:
                _ua = db.query(UserApplication).filter(
                    UserApplication.user_id == current_user.id,
                    UserApplication.application_id == _app.id,
                ).first()
                if not _ua:
                    raise _HTTPEx(
                        status_code=403,
                        detail=f"U heeft geen toegang tot '{_app.name}'. Neem contact op met uw organisatiebeheerder.",
                    )

    parsed = []
    truncation_warnings = []
    for upload in files:
        content = await upload.read()
        ext = upload.filename.split(".")[-1].lower()
        if ext not in ("csv", "xlsx", "xls", "xml", "json"):
            raise HTTPException(400, f"Unsupported file type: {ext}")
        try:
            headers, rows, total = parse_upload(content, upload.filename, ext)
        except ValueError as _perr:
            # Ongeldige JSON/XML (bijv. AFAS-getal met voorloopnul zoals `00`) →
            # duidelijke per-bestand-melding i.p.v. een stille 500.
            raise HTTPException(400, f"Kon '{upload.filename}' niet lezen: {_perr}")
        rows = rows[:MAX_ROWS]   # harde bovengrens
        if total > len(rows):
            truncation_warnings.append({
                "filename": upload.filename,
                "verwerkt": len(rows),
                "totaal":   total,
            })
        parsed.append({"filename": upload.filename, "headers": headers, "rows": rows})

    # Bewaar fase-1 data voor een eventuele benchmark (overschrijft de vorige = cache wissen bij nieuwe scan)
    try:
        _ukey = str(current_user.id) if current_user else None
        if _ukey:
            run_cache.set_current(_ukey, source,
                [{"filename": p["filename"], "headers": p["headers"], "rows": p["rows"]} for p in parsed])
    except Exception:
        pass

    # Vangnet: herkent de gekozen route geen enkel bestand, val terug op de juiste validator
    def _count_recognized(std):
        if std == "zib":
            return sum(1 for p in parsed if detect_zib_schema(p["filename"]))
        if std == "kikv":
            return sum(1 for p in parsed if detect_schema(p["filename"], p["headers"]))
        return sum(1 for p in parsed if _detect_template(p["filename"], p["headers"]))
    if parsed and _count_recognized(standard) == 0:
        for _alt in ("zib", "kikv", "algemeen"):
            if _alt != standard and _count_recognized(_alt) > 0:
                standard = _alt
                break

    # ── Actualiteit (altijd, voor alle bestanden) ─────────────────────────────
    actuality_results = []
    for p in parsed:
        fields = detect_date_fields(p["headers"])
        sk = detect_schema(p["filename"], p["headers"])
        ar = validate_actuality(p["rows"], field_map=fields, max_age_days=max_age_days or 30)
        ar["filename"]   = p["filename"]
        ar["schema_key"] = sk
        # KIK-V actualiteitsnorm op basis van gedetecteerd schema
        ar["kikv_norm"]  = get_kikv_norm_for_schema(sk) if sk else None
        # Voeg source_file toe aan elk individueel issue-item
        for issue_list in ("outdated", "inconsistent"):
            for item in ar.get(issue_list, []):
                item["source_file"] = p["filename"]
        actuality_results.append(ar)

    # ── Algemeen pad ─────────────────────────────────────────────────────────
    if standard == "algemeen":
        alg_input = [{"filename": p["filename"], "headers": p["headers"], "rows": p["rows"]} for p in parsed]
        result = validate_algemeen(alg_input)
        # Cross-file-checks ("personen niet in Medewerker", "overlappende periodes", …)
        # ook op de algemeen/AFAS-route — herstelt de gap-analyse voor AFAS-bron (CSV+JSON).
        try:
            result["cross_checks"] = cross_checks_from_files(alg_input)
        except Exception as _cc_err:
            print(f"[WARN] cross-checks failed (result still returned): {_cc_err}")
            result["cross_checks"] = []
        # Cross-check-bevindingen meetellen in de kop (fouten/waarschuwingen) en in
        # de totaalscore: vindt de gap-analyse iets, dan is het niet '100/foutloos'.
        # De per-bestand-scores blijven ongemoeid (cross-checks zijn niet aan één
        # bestand toe te wijzen).
        _cc = result.get("cross_checks", []) or []
        if _cc:
            _sum = result.setdefault("summary", {})
            _sum["error_count"] = _sum.get("error_count", 0) + sum(1 for c in _cc if c.get("severity") == "error")
            _sum["warn_count"]  = _sum.get("warn_count", 0)  + sum(1 for c in _cc if c.get("severity") == "warning")
            if _sum.get("quality", 0) == 100:
                _sum["quality"] = 99
                _sum["rhadix_index"] = round(_sum.get("completeness", 100) * 99 / 100)
        # Benchmark tegen het AFAS-referentieontwerp (alleen AFAS-bestanden tellen mee)
        try:
            benchmark = benchmark_against_reference(alg_input)
        except Exception as _bench_err:
            print(f"[WARN] benchmark failed (result still returned): {_bench_err}")
            benchmark = {"applicable": False, "error": str(_bench_err)}
        run_id = None
        created_at = None
        try:
            _app_id, _lic_id = _resolve_app_and_license("algemeen", current_user, db)
            run = ValidationRun(
                label=label or f"Algemeen-scan {len(parsed)} bestand{'en' if len(parsed) > 1 else ''}",
                files=[{"filename": p["filename"], "schema_key": "algemeen", "rows": len(p["rows"])} for p in parsed],
                results=result,
                total_rows=sum(len(p["rows"]) for p in parsed),
                error_count=result["summary"]["error_count"],
                warn_count=result["summary"]["warn_count"],
                standard="algemeen",
                tenant_id=current_user.tenant_id if current_user else None,
                created_by=current_user.id if current_user else None,
                application_id=_app_id,
                license_id=_lic_id,
            )
            db.add(run); db.commit(); db.refresh(run)
            run_id = run.id
            created_at = str(run.created_at)
        except Exception:
            pass
        return {**result, "truncation": truncation_warnings, "source": source, "standard": standard, "run_id": run_id, "created_at": created_at,
                "label": label or "Algemeen-scan", "benchmark": benchmark}

    # ── ZIB-pad ───────────────────────────────────────────────────────────────
    if standard == "zib":
        zib_input = [{"filename": p["filename"], "rows": p["rows"]} for p in parsed]
        result = validate_zib(zib_input)

        files_summary = [
            {
                "filename":   p["filename"],
                "schema_key": detect_zib_schema(p["filename"]) or "onbekend",
                "rows":       len(p["rows"]),
            }
            for p in parsed
        ]
        run_id = None
        created_at = None
        try:
            _app_id, _lic_id = _resolve_app_and_license("zib", current_user, db)
            run = ValidationRun(
                label=label or f"ZIB-scan {len(parsed)} bestand{'en' if len(parsed) > 1 else ''}",
                files=files_summary,
                results=result,
                total_rows=sum(len(p["rows"]) for p in parsed),
                error_count=sum(
                    len([i for i in fr.get("issues", []) if i.get("severity") == "error"])
                    for fr in result["file_results"]
                ),
                warn_count=sum(
                    len([i for i in fr.get("issues", []) if i.get("severity") == "warning"])
                    for fr in result["file_results"]
                ),
                score=result["score"],
                status="completed",
                standard="zib",
                tenant_id=current_user.tenant_id if current_user else None,
                created_by=current_user.id if current_user else None,
                application_id=_app_id,
                license_id=_lic_id,
                # Phase 3: ZIB has a single composite score; map to use_case_score
                use_case_score=result.get("score"),
                source_system=parsed[0]["filename"] if parsed else None,
            )
            db.add(run)
            db.commit()
            db.refresh(run)
            run_id = run.id
            created_at = run.created_at.isoformat()
        except Exception as db_err:
            print(f"[WARN] DB save failed (result still returned): {db_err}")
            try:
                db.rollback()
            except Exception:
                pass

        # Verrijk ZIB issues met traceervelden
        for fr in result.get("file_results", []):
            fname  = fr.get("filename", "")
            sk     = fr.get("schema_key") or "onbekend"
            layer  = "zib_availability"
            prescan_layer = "prescan"
            for issue in fr.get("issues", []):
                if issue.get("prescan"):
                    enrich_file_result({"issues": [issue]}, layer=prescan_layer, source_file=fname, schema_key=sk)
                else:
                    enrich_file_result({"issues": [issue]}, layer="zib_quality", source_file=fname, schema_key=sk)

        all_issues = collect_all_issues(zib_result=result, actuality_results=actuality_results)

        return {
            **result,
        "truncation": truncation_warnings, "source": source, "standard": standard,
            "run_id":      run_id,
            "created_at":  created_at,
            "concept_mapping": [],
            "actuality":   actuality_results,
            "all_issues":  all_issues,
        }

    # ── KIK-V-pad ─────────────────────────────────────────────────────────────
    files_input = []
    for p in parsed:
        sk = detect_schema(p["filename"], p["headers"])
        files_input.append({
            "filename":   p["filename"],
            "schema_key": sk,
            "headers":    p["headers"],
            "rows":       p["rows"],
        })

    result = validate_files(files_input)

    run_id = None
    created_at = None
    try:
        _app_id, _lic_id = _resolve_app_and_license("kikv", current_user, db)
        run = ValidationRun(
            label=label or f"Validatie {len(files)} bestand{'en' if len(files) > 1 else ''}",
            files=result["files_summary"],
            results=result,
            total_rows=result["total_rows"],
            error_count=result["total_errors"],
            warn_count=result["total_warns"],
            score=result["score"],
            status="completed",
            standard="kikv",
            tenant_id=current_user.tenant_id if current_user else None,
            created_by=current_user.id if current_user else None,
            application_id=_app_id,
            license_id=_lic_id,
            # Phase 3 subscores — filled after OWL/SPARQL validators run below
            structural_score=None,
            relational_score=None,
            use_case_score=None,
            source_system=parsed[0]["filename"] if parsed else None,
        )
        db.add(run)
        db.commit()
        db.refresh(run)
        run_id = run.id
        created_at = run.created_at.isoformat()
    except Exception as db_err:
        print(f"[WARN] DB save failed (result still returned): {db_err}")
        try:
            db.rollback()
        except Exception:
            pass

    # Stap 2: concept-mapping
    concept_results = []
    for f in files_input:
        sk = f["schema_key"]
        if not sk:
            continue
        aliases = KIKV_REFERENCE.get(sk, {}).get("col_aliases", {})
        fmap    = auto_map(f["headers"], aliases)
        cr      = validate_concept_mapping(sk, f["rows"], fmap)
        # Transformeer naar het formaat dat de frontend verwacht
        summary = cr.get("summary", {})
        schema_label = KIKV_REFERENCE.get(sk, {}).get("label", sk.capitalize())
        total    = summary.get("total_fields", 0)
        matched  = summary.get("mapped_fields", 0)
        # Groepeer velden per ontologie-domein (module)
        domains_map: dict = {}
        for field, fdata in cr.get("fields", {}).items():
            uri = fdata.get("concept_uri", "")
            # Extraheer module uit URI (bijv. onz-pers, onz-g)
            module = uri.split("#")[0].split("/")[-1] if "#" in uri else "onbekend"
            if module not in domains_map:
                domains_map[module] = {"name": module, "fields": [], "all_mapped": True}
            domains_map[module]["fields"].append(field)
            if fdata.get("unmapped_rows", 0) > 0:
                domains_map[module]["all_mapped"] = False
        domains = [
            {"name": d["name"], "status": "volledig" if d["all_mapped"] else "gedeeltelijk"}
            for d in domains_map.values()
        ]
        # Skip schemas without ontologie-mapping (bijv. vestiging/client zonder FIELD_RULES)
        if total == 0:
            continue
        concept_results.append({
            # "schema" en "summary" zijn de veldnamen die ConceptMappingRapport.jsx verwacht
            "schema":       sk,
            "schema_key":   sk,          # bewaard voor achterwaartse compatibiliteit
            "schema_label": schema_label,
            "fields":       cr.get("fields", {}),
            "domains":      domains,
            "summary": {
                "mapping_score": summary.get("mapping_score", 100.0),
                "total_fields":  total,    # was: total_concepts
                "mapped_fields": matched,  # was: matched_concepts
            },
            # Vlak niveau bewaard voor Dashboard.jsx (gebruikt optional chaining)
            "mapping_score":    summary.get("mapping_score", 100.0),
            "total_concepts":   total,
            "matched_concepts": matched,
        })

    # ── OWL structural + relational + SPARQL use-case scores ────────────────
    owl_files_data = [
        {
            "filename":   f["filename"],
            "schema_key": f.get("schema_key") or f.get("schema", ""),
            "rows":       f.get("rows", []),
            "mapping":    auto_map(f["headers"], KIKV_REFERENCE.get(f.get("schema_key") or f.get("schema", ""), {}).get("col_aliases", {})),
        }
        for f in files_input
    ]
    structural_result  = validate_structural(owl_files_data)
    relational_result  = validate_relational(owl_files_data)
    use_case_result    = validate_use_cases(owl_files_data)

    # ── Phase 3: backfill subscores on the saved run ──────────────────────────
    if run_id is not None:
        try:
            _saved_run = db.query(ValidationRun).filter(ValidationRun.id == run_id).first()
            if _saved_run:
                _saved_run.structural_score = structural_result.get("structural_score")
                _saved_run.relational_score = relational_result.get("relational_score")
                _saved_run.use_case_score   = use_case_result.get("use_case_score")
                db.commit()
        except Exception as _e:
            print(f"[WARN] subscore backfill failed: {_e}")
            try:
                db.rollback()
            except Exception:
                pass

    # Verrijk KIK-V issues met traceervelden
    for fsum in result.get("files_summary", []):
        fname = fsum.get("filename", "")
        sk    = fsum.get("schema_key") or ""
        for issue in fsum.get("issues", []):
            layer = "prescan" if issue.get("prescan") else "quality"
            enrich_file_result({"issues": [issue]}, layer=layer, source_file=fname, schema_key=sk)

    all_issues = collect_all_issues(kikv_result=result, actuality_results=actuality_results)

    return {
        **result,
        "truncation": truncation_warnings, "source": source, "standard": standard,
        "run_id":          run_id,
        "created_at":      created_at,
        "concept_mapping": concept_results,
        "actuality":       actuality_results,
        "all_issues":      all_issues,
        # ── Nieuwe scores (Part 1 + 2 + 3 — los van Rhadix Index) ──────────
        "structural_score":    structural_result.get("structural_score"),
        "structural_issues":   structural_result.get("structural_issues", []),
        "structural_checks":   {
            "total":  structural_result.get("checks_total"),
            "passed": structural_result.get("checks_passed"),
            "coverage": structural_result.get("rule_coverage", []),
        },
        "relational_score":    relational_result.get("relational_score"),
        "relational_issues":   relational_result.get("relational_issues", []),
        "relational_fk":       relational_result.get("fk_results", []),
        "use_case_score":      use_case_result.get("use_case_score"),
        "indicator_results":   use_case_result.get("indicator_results", []),
        "rdf_graph_triples":   use_case_result.get("graph_triples", 0),
    }


# AFAS/ONS-template (fase 1) → KIK-V-schema (fase 2). Vangnet zodat een bestand
# dat in fase 1 al herkend is niet stilletjes uit de KIK-V-benchmark valt puur
# omdat de bestandsnaam de KIK-V-trefwoorden mist.
_TEMPLATE_TO_KIKV = {
    "employees":     "medewerker",     "ons_employees": "medewerker",
    "timetable":     "werkovereenkomst", "ons_contracts": "werkovereenkomst",
    "functions":     "functie",
    "illness":       "verzuim",        "ons_absence":   "verzuim",
    "ons_teams":     "vestiging",
}


def _kikv_schema_for(filename: str, headers: list) -> Optional[str]:
    """KIK-V-schema voor een bestand: eerst de KIK-V-detector, anders de
    AFAS/ONS-template-herkenning uit fase 1 doorvertalen."""
    sk = detect_schema(filename, headers)
    if sk:
        return sk
    return _TEMPLATE_TO_KIKV.get(_detect_template(filename, headers))


@router.post("/benchmark")
async def benchmark(
    standard: str = Form(...),                 # "kikv" | "zib"
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_optional_user),
):
    """Fase 2 — benchmark de (gecachte) fase-1 data tegen een gekozen standaard.

    Geen her-upload nodig: de genormaliseerde fase-1 data is bewaard bij de
    gebruiker (run_cache). Draait het gekozen profiel (KIK-V of ZIB) en geeft
    het conformiteitsresultaat terug.
    """
    ukey = str(current_user.id) if current_user else None
    cached = run_cache.get_current(ukey) if ukey else None
    if not cached or not cached.get("files"):
        raise HTTPException(400, "Geen fase-1 data gevonden — voer eerst een validatie uit.")

    files = cached["files"]
    std = (standard or "").lower().strip()

    if std == "zib":
        result = validate_zib([{"filename": f["filename"], "rows": f["rows"]} for f in files])
        return {**result, "benchmark_standard": "zib", "source": cached.get("source")}

    if std == "kikv":
        files_input = [{
            "filename":   f["filename"],
            "schema_key": _kikv_schema_for(f["filename"], f["headers"]),
            "headers":    f["headers"],
            "rows":       f["rows"],
        } for f in files]
        result = validate_files(files_input)
        # Geen enkel bestand aan een KIK-V-schema gekoppeld? Geef dat expliciet terug
        # i.p.v. een misleidende lege score van 100 — met de bestandsnamen erbij.
        if not result.get("file_results"):
            return {
                **result,
                "benchmark_standard": "kikv",
                "source":    cached.get("source"),
                "recognized": False,
                "uploaded_files": [f["filename"] for f in files],
                "note": ("Geen van de aangeleverde bestanden kon aan een KIK-V-schema "
                         "gekoppeld worden. Controleer de bestands- of kolomnamen "
                         "(bijv. medewerker/employees, werkovereenkomst/contract, functie, verzuim)."),
            }
        return {**result, "benchmark_standard": "kikv", "source": cached.get("source"), "recognized": True}

    raise HTTPException(400, f"Onbekende benchmark-standaard: {standard!r}")
